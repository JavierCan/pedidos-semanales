from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
import json
import re
from typing import Any

from openpyxl import load_workbook

from .config import (
    CATALOGOS_DIR,
    COLABORADORES_FILE,
    DESTINATARIOS_FILE,
    OUTPUT_DIR,
    TEMPLATES_DIR,
    TIPOS_FILE,
)


EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

CONFIG_ENVIO_FILE = CATALOGOS_DIR / "config_envio.json"

TEMPLATE_FILENAME = "PEDIDOS_CORPORATIVO.xlsx"
TEMPLATE_SHEET_NAME = "Formato_corporativo"

DEFAULT_TIPOS = [
    {
        "clave": "rejas_30_normal",
        "nombre": "Num. Rejas de 30 piezas (normal)",
        "maximo_individual": 3,
        "activo": True,
    },
    {
        "clave": "rejas_30_jumbo",
        "nombre": "Num. Rejas de 30 piezas (jumbo)",
        "maximo_individual": 3,
        "activo": True,
    },
    {
        "clave": "rejas_18_normal",
        "nombre": "Num. Rejas de 18 piezas (normal)",
        "maximo_individual": 3,
        "activo": True,
    },
]

DEFAULT_CONFIG_ENVIO = {
    "destinatario_principal_id": None,
    "copias_destinatarios_ids": [],
    "asunto_base": "Pedido semanal de huevo",
    "mensaje_base": "Buen día,\n\nComparto el archivo semanal de pedidos de huevo.\n",
}


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_int(value: Any) -> int:
    if value in (None, "", " "):
        return 0
    return int(value)


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return deepcopy(default)

    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return deepcopy(default)

    return json.loads(raw)


def _write_json(path: Path, payload: Any) -> None:
    _ensure_parent(path)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def is_valid_email(email: str) -> bool:
    email = _safe_str(email)
    return bool(email) and bool(EMAIL_REGEX.match(email))


def _validate_optional_email(email: str) -> str:
    email = _safe_str(email)
    if email and not is_valid_email(email):
        raise ValueError("El correo no es válido.")
    return email


def _sort_by_nombre(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(items, key=lambda x: _safe_str(x.get("nombre")).lower())


def _next_id(items: list[dict[str, Any]]) -> int:
    ids: list[int] = []
    for item in items:
        try:
            ids.append(int(item.get("id", 0)))
        except (TypeError, ValueError):
            continue
    return (max(ids) if ids else 0) + 1


def _normalize_colaborador(item: dict[str, Any], fallback_id: int | None = None) -> dict[str, Any]:
    return {
        "id": int(item.get("id", fallback_id or 0)),
        "nombre": _safe_str(item.get("nombre")),
        "area": _safe_str(item.get("area")),
        "correo": _safe_str(item.get("correo")),
        "activo": bool(item.get("activo", True)),
    }


def _normalize_destinatario(item: dict[str, Any], fallback_id: int | None = None) -> dict[str, Any]:
    return {
        "id": int(item.get("id", fallback_id or 0)),
        "nombre": _safe_str(item.get("nombre")),
        "correo": _safe_str(item.get("correo")),
        "activo": bool(item.get("activo", True)),
    }


def _ensure_catalog_ids(path: Path, normalizer, default: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items = _read_json(path, default)
    changed = False
    normalized_items: list[dict[str, Any]] = []

    for idx, item in enumerate(items, start=1):
        normalized = normalizer(item, fallback_id=idx)
        if item != normalized:
            changed = True
        normalized_items.append(normalized)

    if changed or not path.exists():
        _write_json(path, normalized_items)

    return normalized_items


# -----------------------------
# Catálogo de colaboradores
# -----------------------------
def load_colaboradores() -> list[dict[str, Any]]:
    return _sort_by_nombre(_ensure_catalog_ids(COLABORADORES_FILE, _normalize_colaborador, []))


def get_colaborador_by_id(colaborador_id: int) -> dict[str, Any] | None:
    for item in load_colaboradores():
        if int(item["id"]) == int(colaborador_id):
            return deepcopy(item)
    return None


def get_colaboradores_with_email() -> list[dict[str, Any]]:
    return [
        deepcopy(item)
        for item in load_colaboradores()
        if item.get("activo", True) and _safe_str(item.get("correo"))
    ]


def add_colaborador(
    nombre: str,
    area: str,
    correo: str = "",
    activo: bool = True,
) -> dict[str, Any]:
    nombre = _safe_str(nombre)
    area = _safe_str(area)
    correo = _validate_optional_email(correo)

    if not nombre:
        raise ValueError("El nombre del colaborador es obligatorio.")
    if not area:
        raise ValueError("El área es obligatoria.")

    colaboradores = load_colaboradores()

    if any(c["nombre"].lower() == nombre.lower() for c in colaboradores):
        raise ValueError("Ese colaborador ya existe.")

    new_item = {
        "id": _next_id(colaboradores),
        "nombre": nombre,
        "area": area,
        "correo": correo,
        "activo": bool(activo),
    }
    colaboradores.append(new_item)
    _write_json(COLABORADORES_FILE, _sort_by_nombre(colaboradores))
    return new_item


def update_colaborador(
    colaborador_id: int,
    nombre: str,
    area: str,
    correo: str = "",
    activo: bool = True,
) -> dict[str, Any]:
    nombre = _safe_str(nombre)
    area = _safe_str(area)
    correo = _validate_optional_email(correo)

    if not nombre:
        raise ValueError("El nombre del colaborador es obligatorio.")
    if not area:
        raise ValueError("El área es obligatoria.")

    colaboradores = load_colaboradores()
    target = None

    for item in colaboradores:
        if int(item["id"]) == int(colaborador_id):
            target = item
            break

    if target is None:
        raise ValueError("No encontré el colaborador a editar.")

    if any(
        int(item["id"]) != int(colaborador_id) and item["nombre"].lower() == nombre.lower()
        for item in colaboradores
    ):
        raise ValueError("Ya existe otro colaborador con ese nombre.")

    target["nombre"] = nombre
    target["area"] = area
    target["correo"] = correo
    target["activo"] = bool(activo)

    _write_json(COLABORADORES_FILE, _sort_by_nombre(colaboradores))
    return target


def delete_colaborador(colaborador_id: int) -> None:
    colaboradores = load_colaboradores()
    filtered = [c for c in colaboradores if int(c["id"]) != int(colaborador_id)]

    if len(filtered) == len(colaboradores):
        raise ValueError("No encontré el colaborador a eliminar.")

    _write_json(COLABORADORES_FILE, _sort_by_nombre(filtered))


# -----------------------------
# Catálogo de destinatarios
# -----------------------------
def load_destinatarios() -> list[dict[str, Any]]:
    return _sort_by_nombre(_ensure_catalog_ids(DESTINATARIOS_FILE, _normalize_destinatario, []))


def get_destinatario_by_id(destinatario_id: int) -> dict[str, Any] | None:
    for item in load_destinatarios():
        if int(item["id"]) == int(destinatario_id):
            return deepcopy(item)
    return None


def add_destinatario(nombre: str, correo: str, activo: bool = True) -> dict[str, Any]:
    nombre = _safe_str(nombre)
    correo = _safe_str(correo)

    if not nombre:
        raise ValueError("El nombre del destinatario es obligatorio.")
    if not is_valid_email(correo):
        raise ValueError("El correo del destinatario no es válido.")

    destinatarios = load_destinatarios()

    if any(d["nombre"].lower() == nombre.lower() for d in destinatarios):
        raise ValueError("Ese destinatario ya existe.")

    new_item = {
        "id": _next_id(destinatarios),
        "nombre": nombre,
        "correo": correo,
        "activo": bool(activo),
    }
    destinatarios.append(new_item)
    _write_json(DESTINATARIOS_FILE, _sort_by_nombre(destinatarios))
    return new_item


def update_destinatario(
    destinatario_id: int,
    nombre: str,
    correo: str,
    activo: bool = True,
) -> dict[str, Any]:
    nombre = _safe_str(nombre)
    correo = _safe_str(correo)

    if not nombre:
        raise ValueError("El nombre del destinatario es obligatorio.")
    if not is_valid_email(correo):
        raise ValueError("El correo del destinatario no es válido.")

    destinatarios = load_destinatarios()
    target = None

    for item in destinatarios:
        if int(item["id"]) == int(destinatario_id):
            target = item
            break

    if target is None:
        raise ValueError("No encontré el destinatario a editar.")

    if any(
        int(item["id"]) != int(destinatario_id) and item["nombre"].lower() == nombre.lower()
        for item in destinatarios
    ):
        raise ValueError("Ya existe otro destinatario con ese nombre.")

    target["nombre"] = nombre
    target["correo"] = correo
    target["activo"] = bool(activo)

    _write_json(DESTINATARIOS_FILE, _sort_by_nombre(destinatarios))
    return target


def delete_destinatario(destinatario_id: int) -> None:
    destinatarios = load_destinatarios()
    filtered = [d for d in destinatarios if int(d["id"]) != int(destinatario_id)]

    if len(filtered) == len(destinatarios):
        raise ValueError("No encontré el destinatario a eliminar.")

    _write_json(DESTINATARIOS_FILE, _sort_by_nombre(filtered))


# -----------------------------
# Tipos de pedido
# -----------------------------
def load_tipos() -> list[dict[str, Any]]:
    tipos = _read_json(TIPOS_FILE, DEFAULT_TIPOS)

    normalized: list[dict[str, Any]] = []
    for item in tipos:
        normalized.append(
            {
                "clave": _safe_str(item.get("clave")),
                "nombre": _safe_str(item.get("nombre")),
                "maximo_individual": int(item.get("maximo_individual", 3)),
                "activo": bool(item.get("activo", True)),
            }
        )

    if not TIPOS_FILE.exists():
        _write_json(TIPOS_FILE, normalized)

    return normalized


# -----------------------------
# Configuración fija de envío
# -----------------------------
def load_config_envio() -> dict[str, Any]:
    raw = _read_json(CONFIG_ENVIO_FILE, DEFAULT_CONFIG_ENVIO)

    # Compatibilidad con versiones anteriores
    legacy_copias_ids = raw.get("copias_ids", [])
    legacy_copias_destinatarios_ids = raw.get("copias_destinatarios_ids", legacy_copias_ids)

    config = {
        "destinatario_principal_id": raw.get("destinatario_principal_id"),
        "copias_destinatarios_ids": legacy_copias_destinatarios_ids,
        "asunto_base": _safe_str(raw.get("asunto_base")) or DEFAULT_CONFIG_ENVIO["asunto_base"],
        "mensaje_base": _safe_str(raw.get("mensaje_base")) or DEFAULT_CONFIG_ENVIO["mensaje_base"],
    }

    if not CONFIG_ENVIO_FILE.exists() or raw != config:
        _write_json(CONFIG_ENVIO_FILE, config)

    return config


def save_config_envio(
    destinatario_principal_id: int | None,
    copias_destinatarios_ids: list[int] | None = None,
    asunto_base: str = "",
    mensaje_base: str = "",
    copias_ids: list[int] | None = None,  # compatibilidad con versiones previas
    copias_colaboradores_ids: list[int] | None = None,  # se ignora ya
) -> dict[str, Any]:
    if copias_destinatarios_ids is None:
        copias_destinatarios_ids = copias_ids or []

    if destinatario_principal_id is not None and get_destinatario_by_id(int(destinatario_principal_id)) is None:
        raise ValueError("El destinatario principal configurado no existe.")

    for copy_id in copias_destinatarios_ids:
        if get_destinatario_by_id(int(copy_id)) is None:
            raise ValueError("Uno de los destinatarios en copia no existe.")

    payload = {
        "destinatario_principal_id": int(destinatario_principal_id) if destinatario_principal_id is not None else None,
        "copias_destinatarios_ids": [int(x) for x in copias_destinatarios_ids],
        "asunto_base": _safe_str(asunto_base) or DEFAULT_CONFIG_ENVIO["asunto_base"],
        "mensaje_base": _safe_str(mensaje_base) or DEFAULT_CONFIG_ENVIO["mensaje_base"],
    }

    _write_json(CONFIG_ENVIO_FILE, payload)
    return payload


def resolve_config_envio() -> dict[str, Any]:
    config = load_config_envio()
    principal = None
    copias_destinatarios: list[dict[str, Any]] = []

    if config.get("destinatario_principal_id") is not None:
        principal = get_destinatario_by_id(int(config["destinatario_principal_id"]))

    for copy_id in config.get("copias_destinatarios_ids", []):
        item = get_destinatario_by_id(int(copy_id))
        if item:
            copias_destinatarios.append(item)

    return {
        "destinatario_principal": principal,
        "copias_destinatarios": copias_destinatarios,
        "copias_colaboradores": [],  # compatibilidad con app anterior
        "copias": copias_destinatarios,  # compatibilidad con app anterior
        "asunto_base": config["asunto_base"],
        "mensaje_base": config["mensaje_base"],
    }


def get_auto_cc_colaboradores_from_lote(lote: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Devuelve colaboradores que deben ir en copia automática:
    - tienen correo
    - total_pedidos > 0
    - no se duplican por email
    """
    result: list[dict[str, Any]] = []
    seen_emails: set[str] = set()

    for item in lote:
        total = calculate_total_pedidos(item)
        if total <= 0:
            continue

        colaborador_id = item.get("colaborador_id")
        if colaborador_id is None:
            continue

        colaborador = get_colaborador_by_id(int(colaborador_id))
        if not colaborador or not colaborador.get("activo", True):
            continue

        correo = _safe_str(colaborador.get("correo")).lower()
        if not correo:
            continue

        if correo in seen_emails:
            continue

        seen_emails.add(correo)
        result.append(deepcopy(colaborador))

    return result


def get_cc_emails_for_lote(lote: list[dict[str, Any]]) -> list[str]:
    """
    Une:
    - copias fijas desde destinatarios
    - copias automáticas desde colaboradores con pedido
    Devuelve correos únicos.
    """
    config_resolved = resolve_config_envio()
    fixed_cc = config_resolved.get("copias_destinatarios", [])
    auto_cc = get_auto_cc_colaboradores_from_lote(lote)

    emails: list[str] = []
    seen: set[str] = set()

    for item in fixed_cc + auto_cc:
        correo = _safe_str(item.get("correo")).lower()
        if correo and correo not in seen:
            seen.add(correo)
            emails.append(correo)

    return emails


# -----------------------------
# Captura semanal en memoria/lote
# -----------------------------
def calculate_total_pedidos(item: dict[str, Any]) -> int:
    return (
        _safe_int(item.get("rejas_30_normal", 0))
        + _safe_int(item.get("rejas_30_jumbo", 0))
        + _safe_int(item.get("rejas_18_normal", 0))
    )


def validate_pedido_item(item: dict[str, Any], max_total: int = 3) -> tuple[bool, str]:
    for key in ("rejas_30_normal", "rejas_30_jumbo", "rejas_18_normal"):
        value = _safe_int(item.get(key, 0))
        if value < 0:
            return False, "No se permiten valores negativos."

    total = calculate_total_pedidos(item)
    if total > max_total:
        return False, f"El total de pedidos no puede ser mayor a {max_total}."

    return True, ""


def upsert_pedido_lote(lote: list[dict[str, Any]], item: dict[str, Any]) -> list[dict[str, Any]]:
    item = deepcopy(item)

    if "colaborador_id" not in item:
        raise ValueError("El pedido debe incluir 'colaborador_id'.")

    ok, message = validate_pedido_item(item)
    if not ok:
        raise ValueError(message)

    item["total_pedidos"] = calculate_total_pedidos(item)

    updated: list[dict[str, Any]] = []
    replaced = False

    for existing in lote:
        if int(existing.get("colaborador_id", 0)) == int(item["colaborador_id"]):
            updated.append(item)
            replaced = True
        else:
            updated.append(deepcopy(existing))

    if not replaced:
        updated.append(item)

    return updated


def remove_pedido_lote(lote: list[dict[str, Any]], colaborador_id: int) -> list[dict[str, Any]]:
    return [deepcopy(x) for x in lote if int(x.get("colaborador_id", 0)) != int(colaborador_id)]


def clear_lote() -> list[dict[str, Any]]:
    return []


def get_resumen_lote(lote: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resumen: list[dict[str, Any]] = []

    for item in lote:
        resumen.append(
            {
                "colaborador_id": int(item.get("colaborador_id", 0)),
                "colaborador": _safe_str(item.get("colaborador")),
                "area": _safe_str(item.get("area")),
                "rejas_30_normal": _safe_int(item.get("rejas_30_normal", 0)),
                "rejas_30_jumbo": _safe_int(item.get("rejas_30_jumbo", 0)),
                "rejas_18_normal": _safe_int(item.get("rejas_18_normal", 0)),
                "total_pedidos": calculate_total_pedidos(item),
            }
        )

    return resumen


# -----------------------------
# Generación del Excel final
# -----------------------------
def get_template_path(filename: str = TEMPLATE_FILENAME) -> Path:
    path = TEMPLATES_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"No encontré el template Excel en: {path}")
    return path


def build_output_filename(fecha_generacion: date | None = None) -> str:
    fecha_generacion = fecha_generacion or date.today()
    timestamp = datetime.now().strftime("%H%M%S")
    return f"PEDIDO_HUEVO_SEMANAL_{fecha_generacion.strftime('%Y-%m-%d')}_{timestamp}.xlsx"


def _pick_sheet(workbook):
    if TEMPLATE_SHEET_NAME in workbook.sheetnames:
        return workbook[TEMPLATE_SHEET_NAME]
    return workbook[workbook.sheetnames[0]]


def _clear_template_rows(ws) -> None:
    for row in range(2, ws.max_row + 1):
        for col in range(1, 7):
            ws.cell(row=row, column=col).value = None


def _filter_items_for_excel(lote: list[dict[str, Any]]) -> list[dict[str, Any]]:
    valid_items: list[dict[str, Any]] = []

    for item in lote:
        total = calculate_total_pedidos(item)
        if total > 0:
            valid_items.append(deepcopy(item))

    return valid_items


def build_excel_final_from_template(
    lote: list[dict[str, Any]],
    fecha_generacion: date | None = None,
    template_filename: str = TEMPLATE_FILENAME,
    output_filename: str | None = None,
) -> Path:
    items = _filter_items_for_excel(lote)
    if not items:
        raise ValueError("No hay pedidos válidos para generar el Excel final.")

    template_path = get_template_path(template_filename)
    wb = load_workbook(template_path)
    ws = _pick_sheet(wb)

    _clear_template_rows(ws)

    for idx, item in enumerate(items, start=2):
        ws.cell(row=idx, column=1).value = idx - 1
        ws.cell(row=idx, column=2).value = _safe_str(item.get("colaborador"))
        ws.cell(row=idx, column=3).value = _safe_str(item.get("area"))
        ws.cell(row=idx, column=4).value = _safe_int(item.get("rejas_30_normal", 0)) or ""
        ws.cell(row=idx, column=5).value = _safe_int(item.get("rejas_30_jumbo", 0)) or ""
        ws.cell(row=idx, column=6).value = _safe_int(item.get("rejas_18_normal", 0)) or ""

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    final_name = output_filename or build_output_filename(fecha_generacion=fecha_generacion)
    output_path = OUTPUT_DIR / final_name
    wb.save(output_path)

    return output_path


def get_excel_preview_rows(lote: list[dict[str, Any]]) -> list[dict[str, Any]]:
    preview_rows: list[dict[str, Any]] = []
    valid_items = _filter_items_for_excel(lote)

    for idx, item in enumerate(valid_items, start=1):
        preview_rows.append(
            {
                "No.Pedido": idx,
                "Nombre completo del colaborador": _safe_str(item.get("colaborador")),
                "Área": _safe_str(item.get("area")),
                "Num. Rejas de 30 piezas (normal)": _safe_int(item.get("rejas_30_normal", 0)),
                "Num. Rejas de 30 piezas (jumbo)": _safe_int(item.get("rejas_30_jumbo", 0)),
                "Num. Rejas de 18 piezas (normal)": _safe_int(item.get("rejas_18_normal", 0)),
            }
        )

    return preview_rows